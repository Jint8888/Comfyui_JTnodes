"""
JTnodes implementation for ComfyUI - Optimized version
"""
import os
import re
import json
import torch
import numpy as np
from pathlib import Path
from PIL import Image
from PIL.PngImagePlugin import PngInfo
import openpyxl
from openpyxl import Workbook
from .LLM_siliconflow import SiliconflowFreeNode

class JTBrightnessNode:
    """
    A basic image processing node that adjusts image brightness.
    
    Attributes:
        RETURN_TYPES (tuple): Defines the output types for the node
        FUNCTION (str): Name of the processing function
        CATEGORY (str): Node category in the UI
    
    Returns:
        tuple: Contains the processed image tensor with adjusted brightness
    """
    
    @classmethod
    def INPUT_TYPES(cls):
        """Define the input types for the node"""
        return {
            "required": {
                "image": ("IMAGE",),
                "brightness": ("FLOAT", {
                    "default": 1.0,
                    "min": 0.0,
                    "max": 2.0,
                    "step": 0.1
                }),
            },
        }
    
    RETURN_TYPES = ("IMAGE",)
    FUNCTION = "process_image"
    CATEGORY = "JT/image"

    def process_image(self, image: torch.Tensor, brightness: float) -> tuple[torch.Tensor]:
        """
        Adjust the brightness of the input image.
        
        Args:
            image: Input image tensor of shape (B, H, W, C)
            brightness: Brightness adjustment factor (float between 0.0 and 2.0)
        
        Returns:
            tuple: Contains the processed image tensor with adjusted brightness
            
        Raises:
            ValueError: If image is not a torch.Tensor
        """
        # 验证输入
        if not isinstance(image, torch.Tensor):
            raise ValueError("Expected image to be a torch.Tensor")
        
        # 应用亮度调整并限制在有效范围内
        # 使用torch内联操作提高性能
        return (torch.clamp(image * float(brightness), 0.0, 1.0),)

class JTImagesavetopath:
    """
    Enhanced image saver with flexible naming options
    """
    
    @classmethod
    def INPUT_TYPES(cls):
        """Define the input types for the node"""
        return {
            "required": {
                "image": ("IMAGE",),
                "folder_path": ("STRING", {
                    "default": "/path",
                    "multiline": False
                }),
                "filename": ("STRING", {
                    "default": "Image",
                    "multiline": False
                }),
                "format": (["PNG", "JPG"], {
                    "default": "JPG"
                }),
                "use_counter": ("BOOLEAN", {
                    "default": False,
                    "label": "使用序号"
                }),
                "separator": (["none", "hyphen", "underscore"], {
                    "default": "hyphen",
                    "label": "分隔符"
                }),
                "counter_digits": ("INT", {
                    "default": 4,
                    "min": 1,
                    "max": 5,
                    "step": 1,
                    "label": "序号位数(1-5)"
                }),
                "allow_overwrite": ("BOOLEAN", {
                    "default": True,
                    "label": "允许覆盖"
                }),
            },
            "hidden": {
                "prompt": "PROMPT",
                "extra_pnginfo": "EXTRA_PNGINFO"
            },
        }
    
    RETURN_TYPES = ("IMAGE", "STRING", "STRING", "INT")
    RETURN_NAMES = ("image", "save_folder", "save_filename", "save_count")
    FUNCTION = "save_image"
    CATEGORY = "JT/image"

    def _get_save_path(self, folder_path: Path, filename: str, extension: str,
                      use_counter: bool, separator_type: str, digits: int,
                      allow_overwrite: bool, index: int = 0) -> Path:
        """生成保存路径"""
        # 处理基本文件路径（不带序号）
        base_path = folder_path / f"{filename}{extension}"
        if not use_counter:
            return None if not allow_overwrite and base_path.exists() else base_path
            
        # 获取分隔符并处理序号位数
        separator = {"none": "", "hyphen": "-", "underscore": "_"}[separator_type]
        digits = min(max(digits, 1), 5)  # 限制在1-5位之间
        current_number = index + 1

        # 生成带序号的文件名
        def get_path(num): 
            return folder_path / f"{filename}{separator}{num:0{digits}d}{extension}"
            
        # 允许覆盖时直接使用当前序号
        if allow_overwrite:
            return get_path(current_number)
            
        # 不允许覆盖时查找可用序号
        save_path = get_path(current_number)
        while save_path.exists():
            current_number += 1
            save_path = get_path(current_number)
            
        return save_path
    
    def save_image(self, image, folder_path, filename, format, use_counter, 
                  separator, counter_digits, allow_overwrite, prompt=None, extra_pnginfo=None):
        """Save images with advanced naming options"""
        # 输入验证
        if not isinstance(image, torch.Tensor):
            raise ValueError("Expected image to be a torch.Tensor")
        if not (filename := filename.strip()):
            raise ValueError("Filename cannot be empty")
        
        # 初始化保存环境
        folder_path = Path(folder_path)
        folder_path.mkdir(parents=True, exist_ok=True)
        extension = ".png" if format == "PNG" else ".jpg"
        
        # 处理图像数据
        images = image.cpu().numpy()
        if len(images.shape) == 3:
            images = images[np.newaxis, ...]
            
        saved_paths = []
        
        # 生成文件路径并保存图片
        for idx in range(images.shape[0]):
            save_path = self._get_save_path(
                folder_path, filename, extension,
                use_counter, separator, counter_digits,
                allow_overwrite, idx
            )
            
            if save_path is None:  # 不允许覆盖且文件存在
                continue
                
            # 转换并保存图片
            # 转换图像数据
            i = images[idx]
            if len(i.shape) == 3 and i.shape[2] == 4:  # 带有alpha通道
                # 分别处理RGB和alpha通道
                rgb = (i[:, :, :3] * 255).clip(0, 255).astype(np.uint8)
                alpha = (i[:, :, 3] * 255).clip(0, 255).astype(np.uint8)
                # 合并通道
                rgba = np.dstack((rgb, alpha))
                img = Image.fromarray(rgba, mode='RGBA')
            else:  # 普通RGB图像
                rgb = (i * 255).clip(0, 255).astype(np.uint8)
                img = Image.fromarray(rgb, mode='RGB')
            
            # 设置保存参数
            save_params = {}
            if format == "PNG":
                save_params["format"] = "PNG"
                if img.mode == 'RGBA':
                    save_params["optimize"] = False  # 避免优化影响alpha通道
                    save_params["compress_level"] = 1  # 使用较低压缩率
            else:
                save_params["format"] = "JPEG"
                save_params["quality"] = 95
            
            # 添加PNG元数据
            if format == "PNG" and (prompt or extra_pnginfo):
                pnginfo = PngInfo()
                if prompt:
                    pnginfo.add_text("prompt", json.dumps(prompt))
                if extra_pnginfo:
                    pnginfo.add_text("workflow", json.dumps(extra_pnginfo))
                save_params["pnginfo"] = pnginfo
                
            img.save(save_path, **save_params)
            saved_paths.append(save_path)
        
        # 返回结果
        return (
            image,                          # 原始图像
            str(folder_path.absolute()),    # 保存目录
            "\n".join(p.name for p in saved_paths) if saved_paths else "",  # 文件名列表
            len(saved_paths)                # 保存数量
        )

class JTcounter:
    """
    A number sequence generator that converts integers into formatted serial numbers.
    
    Attributes:
        RETURN_TYPES (tuple): Defines the output type as STRING
        FUNCTION (str): Name of the processing function
        CATEGORY (str): Node category in the UI
        
    Features:
        - Supports 1-5 digit serial numbers
        - Automatic handling of number overflow
        - Smart digit padding based on input
    """
    
    @classmethod
    def INPUT_TYPES(cls):
        """定义输入参数"""
        return {
            "required": {
                "number": ("INT", {
                    "default": 1,
                    "min": 0,
                    "max": 99999,
                    "step": 1,
                    "display": "number"
                }),
                "digits": ("INT", {
                    "default": 4,
                    "min": 1,
                    "max": 5,
                    "step": 1,
                    "display": "number",
                    "label": "序列号位数(1-5)"
                }),
            },
        }
    
    RETURN_TYPES = ("STRING",)
    FUNCTION = "generate_serial"
    CATEGORY = "JT/text"

    def generate_serial(self, number: int, digits: int) -> tuple[str]:
        """
        Generate a formatted serial number string.
        
        Args:
            number: Input number to convert
            digits: Number of digits for the serial (1-5)
            
        Returns:
            tuple: Contains the formatted serial number string
        """
        # 验证并限制位数范围(1-5)
        digits = min(max(digits, 1), 5)
        
        # 计算数字实际需要的位数
        required_digits = len(str(number)) if number > 0 else 1
        
        # 确定最终位数(不超过5位)
        final_digits = min(max(digits, required_digits), 5)
        
        # 格式化为指定位数的字符串
        return (f"{number:0{final_digits}d}",)

class JTSaveTextToFile:
    """文本文件保存节点，支持追加和覆盖模式
    
    Attributes:
        RETURN_TYPES (tuple): 定义输出类型为STRING
        FUNCTION (str): 处理函数名
        CATEGORY (str): 节点分类
    """
    
    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": {
                "text": ("STRING", {
                    "default": "",
                    "multiline": True
                }),
                "folder_path": ("STRING", {
                    "default": "/path",
                    "multiline": False
                }),
                "filename": ("STRING", {
                    "default": "output.txt",
                    "multiline": False
                }),
                "write_mode": (["append", "overwrite"], {
                    "default": "append",
                    "label": "写入模式"
                }),
            },
        }
    
    RETURN_TYPES = ("STRING",)
    FUNCTION = "save_text"
    CATEGORY = "JT/text"

    def save_text(self, text: str, folder_path: str, filename: str, write_mode: str) -> tuple[str]:
        """保存文本到文件
        
        Args:
            text: 文本内容
            folder_path: 保存目录
            filename: 文件名
            write_mode: 写入模式(append/overwrite)
        """
        # 创建保存目录
        save_path = Path(folder_path)
        save_path.mkdir(parents=True, exist_ok=True)
        
        # 完整文件路径
        file_path = save_path / filename
        
        # 写入模式
        mode = 'a' if write_mode == 'append' else 'w'
        
        # 写入文件
        with open(file_path, mode, encoding='utf-8') as f:
            # 追加模式下，如果文件存在且非空则添加换行
            if mode == 'a' and file_path.exists() and file_path.stat().st_size > 0:
                f.write('\n')
            f.write(text)
            
        return (text,)

class JTSaveTextToExcel:
    """Excel表格保存节点
    
    Attributes:
        RETURN_TYPES (tuple): 定义输出类型为STRING
        FUNCTION (str): 处理函数名
        CATEGORY (str): 节点分类
        
    Notes:
        - 自动处理文件扩展名(.xlsx)
        - 支持指定工作表名和单元格位置
        - 自动创建或更新工作表
        - 如果输入文本包含多行，仅保存第一行
        - 输出实际保存到表格中的内容（第一行文本）
    """
    
    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": {
                "text": ("STRING", {
                    "default": "",
                    "multiline": True
                }),
                "folder_path": ("STRING", {
                    "default": "/path",
                    "multiline": False
                }),
                "filename": ("STRING", {
                    "default": "output",
                    "multiline": False
                }),
                "sheet_name": ("STRING", {
                    "default": "Sheet1",
                    "multiline": False
                }),
                "row": ("INT", {
                    "default": 1,
                    "min": 1,
                    "max": 1048576,  # Excel最大行数
                    "step": 1
                }),
                "column": ("INT", {
                    "default": 1,
                    "min": 1,
                    "max": 16384,  # Excel最大列数
                    "step": 1
                })
            },
        }
    
    RETURN_TYPES = ("STRING",)
    FUNCTION = "save_to_excel"
    CATEGORY = "JT/text"

    def save_to_excel(self, text: str, folder_path: str, filename: str,
                     sheet_name: str, row: int, column: int) -> tuple[str]:
        """保存文本到Excel表格
        
        Args:
            text: 文本内容
            folder_path: 保存目录
            filename: 文件名(可带扩展名)
            sheet_name: 工作表名
            row: 起始行号
            column: 起始列号
        """
        # 创建保存目录
        save_path = Path(folder_path)
        save_path.mkdir(parents=True, exist_ok=True)
        
        # 处理文件名
        try:
            # 处理文件扩展名
            if not any(filename.endswith(ext) for ext in ['.xlsx', '.xls']):
                filename = f"{filename}.xlsx"
            
            # 完整文件路径
            file_path = save_path / filename
            
            # 获取或创建工作簿
            wb = openpyxl.load_workbook(file_path) if file_path.exists() else Workbook()
            
            # 获取或创建工作表
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.create_sheet(sheet_name)
                # 如果是默认的Sheet且不是目标工作表，删除它
                if "Sheet" in wb.sheetnames and sheet_name != "Sheet":
                    wb.remove(wb["Sheet"])
            
            # 处理文本内容（如果有多行，只取第一行）
            first_line = text.split('\n')[0] if text else ""
            
            # 写入文本内容
            ws.cell(row=row, column=column, value=first_line)
            
            # 保存文件
            wb.save(file_path)
            
        except Exception as e:
            raise RuntimeError(f"保存Excel文件时出错: {str(e)}")
        
        # 返回实际保存的内容
        return (first_line,)

class JTFindTextFromExcel:
    """Find specified text in Excel file and return related information
    
    Attributes:
        RETURN_TYPES (tuple): Output types definition
        FUNCTION (str): Processing function name
        CATEGORY (str): Node category in UI
        
    Features:
        - Find specified text in Excel file
        - Return text from specified column in the same row
        - Return row and column numbers of found text
        
    Notes:
        - Automatically handles Excel file extension (.xlsx)
        - Creates directory if not exists
    """
    
    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": {
                "Excel_Filepath": ("STRING", {
                    "default": "/path",
                    "multiline": False,
                    "label": "Excel_Filepath",
                    "paste": True
                }),
                "Excel_Filename": ("STRING", {
                    "default": "table",
                    "multiline": False,
                    "label": "Excel_Filename",
                    "paste": True
                }),
                "Find_Text": ("STRING", {
                    "default": "",
                    "multiline": False,
                    "label": "Find_Text",
                    "paste": True
                }),
                "Output_Column": ("INT", {
                    "default": 1,
                    "min": 1,
                    "max": 16384,  # Excel max column number
                    "step": 1,
                    "label": "Output_Column",
                    "paste": True,
                    "round": True
                }),
            },
        }
    
    RETURN_TYPES = ("STRING", "INT", "INT")
    RETURN_NAMES = ("found_text", "row_number", "column_number")
    FUNCTION = "find_in_excel"
    CATEGORY = "JT/text"

    def find_in_excel(self, Excel_Filepath: str, Excel_Filename: str, Find_Text: str, Output_Column: int) -> tuple[str, int, int]:
        """Find text in Excel file
        
        Args:
            Excel_Filepath: Excel file directory path
            Excel_Filename: Excel file name
            Find_Text: Text to search for
            Output_Column: Column number to output text from
            
        Returns:
            tuple: (found text, row number, column number)
        """
        try:
            # Process file path
            file_dir = Path(Excel_Filepath)
            # Ensure directory exists
            file_dir.mkdir(parents=True, exist_ok=True)
            
            # Handle file extension
            if not any(Excel_Filename.endswith(ext) for ext in ['.xlsx', '.xls']):
                Excel_Filename = f"{Excel_Filename}.xlsx"
                
            # Full file path
            file_path = file_dir / Excel_Filename
            
            # Load Excel file
            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb.active
            
            # Search text in cells
            found_text = ""
            row_number = 0
            column_number = 0
            
            for row in range(1, ws.max_row + 1):
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row, column=col)
                    if cell.value and str(cell.value) == Find_Text:
                        # Get found position
                        row_number = row
                        column_number = col
                        # Get text from output column
                        output_cell = ws.cell(row=row, column=Output_Column)
                        found_text = str(output_cell.value) if output_cell.value is not None else ""
                        break
                if row_number > 0:  # Stop searching if found
                    break
            
            wb.close()
            return (found_text, row_number, column_number)
            
        except Exception as e:
            raise RuntimeError(f"Error processing Excel file: {str(e)}")

class JTReadFromExcel:
    """Read text from specified position in Excel file
    
    Attributes:
        RETURN_TYPES (tuple): Output types definition
        FUNCTION (str): Processing function name
        CATEGORY (str): Node category in UI
        
    Features:
        - Read text from specified row and column
        - Return the text and its position
        
    Notes:
        - Automatically handles Excel file extension (.xlsx)
        - Creates directory if not exists
    """
    
    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": {
                "Excel_Filepath": ("STRING", {
                    "default": "/path",
                    "multiline": False,
                    "label": "Excel_Filepath",
                    "paste": True
                }),
                "Excel_Filename": ("STRING", {
                    "default": "table",
                    "multiline": False,
                    "label": "Excel_Filename",
                    "paste": True
                }),
                "Row_Number": ("INT", {
                    "default": 1,
                    "min": 1,
                    "max": 1048576,  # Excel max row number
                    "step": 1,
                    "label": "Row_Number",
                    "paste": True,
                    "round": True
                }),
                "Column_Number": ("INT", {
                    "default": 1,
                    "min": 1,
                    "max": 16384,  # Excel max column number
                    "step": 1,
                    "label": "Column_Number",
                    "paste": True,
                    "round": True
                })
            },
        }
    
    RETURN_TYPES = ("STRING", "INT", "INT")
    RETURN_NAMES = ("cell_text", "row_number", "column_number")
    FUNCTION = "read_from_excel"
    CATEGORY = "JT/text"

    def read_from_excel(self, Excel_Filepath: str, Excel_Filename: str, Row_Number: int, Column_Number: int) -> tuple[str, int, int]:
        """Read text from specified position in Excel file
        
        Args:
            Excel_Filepath: Excel file directory path
            Excel_Filename: Excel file name
            Row_Number: Row number to read from
            Column_Number: Column number to read from
            
        Returns:
            tuple: (cell text, row number, column number)
        """
        try:
            # Process file path
            file_dir = Path(Excel_Filepath)
            # Ensure directory exists
            file_dir.mkdir(parents=True, exist_ok=True)
            
            # Handle file extension
            if not any(Excel_Filename.endswith(ext) for ext in ['.xlsx', '.xls']):
                Excel_Filename = f"{Excel_Filename}.xlsx"
                
            # Full file path
            file_path = file_dir / Excel_Filename
            
            # Load Excel file
            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb.active
            
            # Check if position is valid
            if Row_Number > ws.max_row:
                raise ValueError(f"Row number {Row_Number} exceeds maximum row {ws.max_row}")
            if Column_Number > ws.max_column:
                raise ValueError(f"Column number {Column_Number} exceeds maximum column {ws.max_column}")
            
            # Get cell value
            cell = ws.cell(row=Row_Number, column=Column_Number)
            cell_text = str(cell.value) if cell.value is not None else ""
            
            wb.close()
            return (cell_text, Row_Number, Column_Number)
            
        except Exception as e:
            raise RuntimeError(f"Error reading Excel file: {str(e)}")

# Node registration mappings
NODE_CLASS_MAPPINGS = {
    "JT Find Text From Excel": JTFindTextFromExcel,
    "JT Read From Excel": JTReadFromExcel,
    "JTBrightness": JTBrightnessNode,
    "JTImagesavetopath": JTImagesavetopath,
    "JTcounter": JTcounter,
    "SiliconflowFree": SiliconflowFreeNode,
    "JTSaveTextToFile": JTSaveTextToFile,
    "JTSaveTextToExcel": JTSaveTextToExcel
}

NODE_DISPLAY_NAME_MAPPINGS = {
    "JT Find Text From Excel": "JT Find Text From Excel",
    "JT Read From Excel": "JT Read From Excel",
    "JTBrightness": "JT Brightness Adjustment",
    "JTImagesavetopath": "JT Save Image to Path",
    "JTcounter": "JT Serial Counter",
    "SiliconflowFree": "JT Siliconflow LLM",
    "JTSaveTextToFile": "JT Save Text to File",
    "JTSaveTextToExcel": "JT Save Text to Excel"
}
